import openpyxl

openpyxl.Workbook()

book = openpyxl.Workbook()

book.create_sheet("Frutas")
print(book.sheetnames)

book["Frutas"]

frutas_page = book["Frutas"]

frutas_page.append(["Frutas", "Quantidade", "Preço"])
frutas_page.append(["melão", "10", "R$ 5,90"])
frutas_page.append(["Maçã", "15", "R$ 6,90"])
frutas_page.append(["Caju", "50", "R$ 1,90"])

book.save("Planilha de compras.xlsx")


