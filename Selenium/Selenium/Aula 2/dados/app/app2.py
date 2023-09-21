import openpyxl
openpyxl.load_workbook("Planilha de compras.xlsx")

book = openpyxl.load_workbook("Planilha de compras.xlsx")

book["Frutas"]

frutas_page = book["Frutas"]

for rows in frutas_page.iter_rows(min_row=2, max_row=4):
    for cell in rows:
        print(cell.value)

for rows in frutas_page.iter_rows(min_row=2,max_row=4):
    print(rows[0].value,rows[1].value,rows[2].value)

for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        if cell.value == "Banana":
            cell.value = "Fruta 01"

book.save("Planilha de compras.xlsx")

book.save("Planilha de compras V2.xlsx")
