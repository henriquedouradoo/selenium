import openpyxl

book = openpyxl.load_workbook("DADOSS01.xlsx")

print(book.sheetnames)


book.create_sheet("DADOS02")
print(book.sheetnames)

DADOS02_page = book["DADOS02"]
DADOS02_page.append(["Salário", "Horas", "Presença"])
DADOS02_page.append(["2000", "10", "Sim"])
DADOS02_page.append(["4000", "15", "Parcial"])
DADOS02_page.append(["3000", "50", "Imparcial"])

book.save("DADOS02.xlsx")